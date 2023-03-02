import json
from datetime import datetime

from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
import time

START_ROW: int = 2
END_ROW: int = 7

FULL_NAME_COLUMN = 'G'
PHONE_NUMBER_COLUMNS = ['H', 'J', 'K']
NUMBER_BEDROOMS_COLUMN = 'P'
MESSAGES_LOGS_COLUMN = 'Q'
# BUILDING_NAME_COLUMN = 'D'
BUILDING_NAME = 'Fountain Views 3'
PROJECT_NAME = 'TarFV3'

receiver_col_index = column_index_from_string(FULL_NAME_COLUMN)
messages_logs_col_ind = column_index_from_string(MESSAGES_LOGS_COLUMN)
# building_name_col = column_index_from_string(BUILDING_NAME_COLUMN)
number_bedroom_col = column_index_from_string(NUMBER_BEDROOMS_COLUMN)

success_counter: int = 0
fail_counter: int = 0
time_start = datetime.now()

with open(f'success_list.json') as success_file:
    success_list: list = json.load(success_file)

with open(f'failed_list.json') as failed_file:
    failed_list: list = json.load(failed_file)


def name_convertor(full_name: str) -> str:
    first_name = full_name.split()[0]
    return first_name.lower().capitalize()


def saver_json_and_exel():
    with open(f'success_list.json', 'w') as file:
        # Write the list to the file in JSON format
        json.dump(success_list, file)

    with open(f'failed_list.json', 'w') as file:
        # Write the list to the file in JSON format
        json.dump(failed_list, file)

    wb.save(f'{PROJECT_NAME}.xlsx')


# Load the Excel sheet containing the receivers' phone numbers
wb = load_workbook(f'{PROJECT_NAME}.xlsx')
ws = wb.active

# Set up the Chrome driver and navigate to WhatsApp Web
driver = webdriver.Chrome()
driver.get('https://web.whatsapp.com/')

# Wait for the user to log in to WhatsApp Web
input('Please log in to WhatsApp Web, then press Enter to continue...')

# Loop through each receiver's phone number and send a message
for row_index, row in enumerate(ws.iter_rows(min_row=START_ROW, max_row=END_ROW, values_only=True), start=START_ROW):
    print(row_index, row)
    logs_cell = ws.cell(row=row_index, column=messages_logs_col_ind)
    for phone_col_letter in PHONE_NUMBER_COLUMNS:

        saver_json_and_exel()

        phone_col_index = column_index_from_string(phone_col_letter)

        # Get the receiver's phone number and message from the Excel sheet
        phone_number = str(row[phone_col_index-1])

        if phone_number == '0':
            continue
        phone_number = '+' + phone_number
        if phone_number in success_list:
            if logs_cell.value is None:
                logs_cell.value = f'Previously successfully sent to {phone_number} on Whatsapp;\n'
            else:
                logs_cell.value += f'Previously successfully sent to {phone_number} on Whatsapp;\n'
            continue
        if phone_number in failed_list:
            if logs_cell.value is None:
                logs_cell.value = f'Previously not sent to {phone_number} on Whatsapp;\n'
            else:
                logs_cell.value += f'Previously not sent to {phone_number} on Whatsapp;\n'
            continue

        receiver = name_convertor(str(row[receiver_col_index-1]))
        number_bedroom = row[number_bedroom_col-1].split()[0]

        # Navigate to the chat with the receiver
        driver.get(f'https://web.whatsapp.com/send?phone={phone_number}')

        # Wait for the chat to load
        time.sleep(10)

        message = (
            f'''Dear {receiver},\nThis is Vlad, I am a downtown specialist. Just want to check if you are still the owner of {number_bedroom} bedroom apartment in {BUILDING_NAME} and I was wondering if you might be interested to sell or lease it. I am dealing with many investors that are interested in your property. Feel free to contact me for any inquiries.\nThank you & best regards,\nVlad'''
        )
        # Type the message and send it

        try:
            input_box = driver.find_element(By.XPATH, '//div[@class="_3Uu1_"]')
            input_box.send_keys(message)
            input_box.send_keys(Keys.RETURN)
        except NoSuchElementException:
            print(f'Failed send to {phone_number} on Whatsapp;\n')
            failed_list.append(phone_number)
            fail_counter += 1
            if logs_cell.value is None:
                logs_cell.value = f'Failed send to {phone_number} on Whatsapp;\n'
            else:
                logs_cell.value += f'Failed send to {phone_number} on Whatsapp;\n'
            continue

        # Wait for the message to be sent
        time.sleep(10)
        success_list.append(phone_number)
        success_counter += 1

        print(f'Message to {phone_number} sent {str(datetime.now())}!\n')
        if logs_cell.value is None:
            logs_cell.value = f'Message to {phone_number} sent {str(datetime.now())}!;\n'
        else:
            logs_cell.value += f'Message to {phone_number} sent {str(datetime.now())}!;\n'


# save json  and exel
saver_json_and_exel()
# Close the Chrome driver
driver.quit()

with open(f'logs_{PROJECT_NAME}.txt', 'a') as logs:
    logs.write(
        f'\n\nFrom rows {START_ROW} to {END_ROW}: {success_counter} messages were sent between {time_start} and {datetime.now()};\n{fail_counter} were failed!'
    )
print(f'{success_counter} messages sent!\n{fail_counter} failed!')
